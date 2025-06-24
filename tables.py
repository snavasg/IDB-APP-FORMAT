import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


def develop_chal_table(wb, *,data: pd.DataFrame,sheet_name: str = "develop_challenge.xlsx") -> str:
    """
    Crea la hoja 'C. Desafío para el desarrollo' con todo el formato
    (celdas combinadas, colores, bordes) y la rellena con los datos
    entregados.

    Parameters
    ----------
    data : pd.DataFrame
        Columnas requeridas:
          • Element type  ('General Objective' | 'Specific Objective' | 'Result indicator')
          • Number        (p. ej. 1.1, 1.1.A)
          • Name          (texto)
    filepath : str, optional
        Ruta donde se guardará el archivo Excel (por defecto 'develop_challenge.xlsx').

    Returns
    -------
    str
        Ruta final del archivo creado.
    """
    # ─────────────────── Preparar datos ────────────────────────────
    gen_obj_name = (
        data.query("`Element type` == 'General Objective'")["Name"].iloc[0]
        if "General Objective" in data["Element type"].values
        else "[Objetivo General]"
    )

    spec_map = {}
    for _, row in data.iterrows():
        etype, num, name = row["Element type"], str(row["Number"]), row["Name"]
        if etype == "Specific Objective":
            spec_map.setdefault(num, {"objective": name, "indicators": []})
        elif etype == "Result indicator":
            prefix = ".".join(num.split(".")[:2])   # 1.1.A → 1.1
            spec_map.setdefault(prefix, {"objective": None, "indicators": []})
            spec_map[prefix]["indicators"].append(name)

    ordered_specs = sorted(
        spec_map,
        key=lambda k: tuple(map(int, k.split(".")))  # orden numérico 1.1, 1.2, 2.1, etc.
    )
    for v in spec_map.values():
        v["indicators"].sort()

    # ─────────────────── Estilos ───────────────────────────────────
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]                 # empieza de cero
    ws = wb.create_sheet(sheet_name)
    ws.title = "C. Desafío"

    for col, width in zip("ABCDEF", [22, 35, 22, 22, 32, 15]):
        ws.column_dimensions[col].width = width

    header_font = Font(bold=True, color="FFFFFF")
    title_font  = Font(bold=True, size=14)
    instr_font  = Font(italic=True, size=11)
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    fill_dark   = PatternFill("solid", fgColor="004B6B")
    fill_cyan   = PatternFill("solid", fgColor="00B0F0")
    fill_gray   = PatternFill("solid", fgColor="BFBFBF")
    thin        = Border(*(Side(style="thin"),) * 4)

    # ─────────────────── Encabezados fijos ─────────────────────────
    ws.merge_cells("A1:F1")
    ws["A1"].value, ws["A1"].font = "C. Desafío para el desarrollo", title_font

    ws.merge_cells("A2:F2")
    ws["A2"].value, ws["A2"].font = (
        "Instrucciones: Complete las columnas B, E y F.",
        instr_font,
    )

    ws.merge_cells("A3:F3")
    ws["A3"].value = (
        "¿Cuál es el principal reto de desarrollo que trata de abordar el proyecto?"
    )
    ws["A3"].font = Font(bold=True)

    headers = [
        ("A4", "Objetivo General", fill_dark),
        (
            "B4",
            "Identificar los supuestos\nprincipales que vinculan los\n"
            "Objetivos Específicos con el\nObjetivo General",
            fill_cyan,
        ),
        ("C4", "Objetivo Específico", fill_dark),
        ("D4", "Indicadores de\nresultado", fill_dark),
        (
            "E4",
            "¿Hay alguna dimensión del\nObjetivo Específico que no tenga\n"
            "un indicador de resultados?\n[SÍ/NO]",
            fill_dark,
        ),
        ("F4", "Explique", fill_dark),
    ]
    for cell, text, bg in headers:
        ws[cell].value = text
        ws[cell].font = header_font
        ws[cell].alignment = center
        ws[cell].fill = bg
        ws[cell].border = thin

    # ─────────────────── Poblado dinámico ──────────────────────────
    row = 5                     # primera fila de datos
    general_start = row

    for spec_key in ordered_specs:
        spec = spec_map[spec_key]
        indicators = spec["indicators"]
        start = row
        end = row + len(indicators) - 1

        # Combinar columnas B, C, E, F para este bloque
        for col in ("B", "C", "E", "F"):
            ws.merge_cells(f"{col}{start}:{col}{end}")
            ws[f"{col}{start}"].alignment = center
            ws[f"{col}{start}"].border = thin

        # Texto del Objetivo Específico
        ws[f"C{start}"].value = (
            spec["objective"] or f"[Objetivo Específico {spec_key}]"
        )

        # Indicadores de resultado
        for i, ind in enumerate(indicators):
            ws[f"D{row + i}"].value = ind
            ws[f"D{row + i}"].alignment = Alignment(vertical="top")
            ws[f"D{row + i}"].border = thin
            # Bordes en B, E, F para cada fila
            for col in ("B", "E", "F"):
                ws[f"{col}{row + i}"].border = thin

        row = end + 1  # SIN fila en blanco entre bloques

    general_end = row - 1
    ws.merge_cells(f"A{general_start}:A{general_end}")
    ws[f"A{general_start}"].value = gen_obj_name
    ws[f"A{general_start}"].alignment = center
    ws[f"A{general_start}"].border = thin
    for r in range(general_start, general_end + 1):
        ws[f"A{r}"].border = thin

    # ─────────────────── Bloque gris para Indicadores GO ────────────
    go_start = general_end + 1
    for r, txt in zip(
        range(go_start, go_start + 3),
        ["Indicador GO 1", "Indicador GO 2", "Indicador GO 3"],
    ):
        for col in "ABCDEF":
            ws[f"{col}{r}"].fill = fill_gray
            ws[f"{col}{r}"].border = thin
        ws[f"A{r}"].value = txt

    return ws


def create_result_measure_table(wb,*,data: pd.DataFrame,sheet_name: str = "result_measure_table.xlsx") -> str:
    # ── ordenar datos ──
    spec = {}
    for _, row in data.iterrows():
        t, num, name = row["Element type"], str(row["Number"]), row["Name"]
        if t == "Specific Objective":
            spec.setdefault(num, {"objective": name, "inds": []})
        elif t == "Result indicator":
            key = ".".join(num.split(".")[:2])
            spec.setdefault(key, {"objective": None, "inds": []})
            spec[key]["inds"].append(name)

    ordered = sorted(spec, key=lambda k: tuple(map(int, k.split("."))))
    for v in spec.values():
        v["inds"].sort()

    # ── libro y estilos ──
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]                 # empieza de cero
    ws = wb.create_sheet(sheet_name)
    ws.title = "E. Medición"
    widths = [25, 22, 18, 15, 15, 15, 10, 30,
              28, 28, 30, 26, 18, 18, 20, 32, 18, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # Colores
    C_A_G   = "196E8C"  # azul encabezado A-G
    C_H     = "FFFFFF"  # blanco
    C_BANDA = "308144"  # verde banda fila 4
    C_IJ    = "E7E6E6"  # gris claro I-J
    C_KR    = "ACCDB4"  # verde claro K-R

    # Fuentes y alineación
    bold_white = Font(bold=True, color="FFFFFF")
    bold       = Font(bold=True)
    title_fnt  = Font(bold=True, size=14)
    ital_fnt   = Font(italic=True, size=11)
    wrap_ctr   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap_top   = Alignment(vertical="top", wrap_text=True)

    thin = Border(*(Side(style="thin"),)*4)   # solo para datos

    # ── filas 1-3: título, nota ──
    ws.merge_cells("A1:R1"); ws["A1"].value, ws["A1"].font = "E. Medición de Resultados", title_fnt
    ws.merge_cells("A2:R2"); ws["A2"].value, ws["A2"].font = (
        "Instrucciones: complete las columnas H-I y K-R para cada Indicador de Resultado.", ital_fnt)
    ws.merge_cells("A3:R3"); ws["A3"].value = (
        "Nota: Los valores corresponden a la Matriz de Resultados reportada en el Plan de Arranque.")

    # ── FILA 4 ──
    ws.merge_cells("I4:R4")
    k4 = ws["I4"]; k4.value, k4.font, k4.alignment, k4.fill = (
        "Medición de Resultados", bold, wrap_ctr, PatternFill("solid", fgColor=C_BANDA))

    for col in ("I", "J"):
        ws[f"{col}5"].fill = PatternFill("solid", fgColor=C_IJ)

    ws["H4"].fill = PatternFill("solid", fgColor=C_H)

    # ── FILAS 5-6 ──
    # H5-H6 pregunta
    ws.merge_cells("H5:H6")
    h = ws["H5"]; h.value = ("¿Cuándo se incluyó el indicador en la Matriz de Resultados?\n"
                             "(seleccione una opción)")
    h.font, h.alignment, h.fill = bold, wrap_ctr, PatternFill("solid", fgColor=C_H)

    # I-J (gris), K-L-R (verde claro)
    vert_hdrs = [
        ("I", C_IJ, "Medios de Verificación\n(información obtenida de Convergencia)"),
        ("J", C_IJ, "Observaciones del Indicador\n(información obtenida de Convergencia)"),
        ("K", C_KR, "Método de Cálculo\n(explique la metodología)"),
        ("L", C_KR, "Método de Atribución\n(ej. Antes y Después, Evaluación de Impacto)"),
        ("R", C_KR, "Otros problemas\no consideraciones de medición"),
    ]
    for col, color, text in vert_hdrs:
        ws.merge_cells(f"{col}5:{col}6")
        c = ws[f"{col}5"]; c.value, c.font, c.alignment, c.fill = text, bold, wrap_ctr, PatternFill("solid", fgColor=color)

    # M5-O5  (grupo existente)
    ws.merge_cells("M5:O5")
    m = ws["M5"]; m.value, m.font, m.alignment, m.fill = (
        "Si la fuente de datos ya existe", bold, wrap_ctr, PatternFill("solid", fgColor=C_KR))

    # P5-Q5 (grupo recolección)
    ws.merge_cells("P5:Q5")
    p = ws["P5"]; p.value, p.font, p.alignment, p.fill = (
        "Si se recopilarán datos para la evaluación", bold, wrap_ctr, PatternFill("solid", fgColor=C_KR))

    # Sub-encabezados fila 6 M-Q
    subs = {
        "M6": "Fuente de Datos\n(definir la fuente de datos)",
        "N6": "Acceso a Datos\n(definir el proceso para acceder a los datos)",
        "O6": "Periodicidad de los Datos\n(¿Está alineada a la medición?)",
        "P6": "Plan de recolección de Datos\n(plan y consideraciones de tiempos)",
        "Q6": "Responsable\n(definir quién es responsable)",
    }
    for cell, txt in subs.items():
        c = ws[cell]; c.value, c.font, c.alignment, c.fill = txt, bold, wrap_ctr, PatternFill("solid", fgColor=C_KR)

    # Encabezado azul A-G fila 6
    left = ["Objetivos Específicos", "Indicador de Resultado", "Desagregación",
            "Unidad de Medida", "Línea de Base", "Año Línea de Base", "Meta"]
    for i, txt in enumerate(left, 1):
        c = ws.cell(6, i, txt)
        c.font, c.alignment, c.fill = bold_white, wrap_ctr, PatternFill("solid", fgColor=C_A_G)

    # ── DATOS (fila 7+) ──
    row = 7
    for key in ordered:
        block, inds = spec[key], spec[key]["inds"]
        start, end = row, row + len(inds) - 1
        ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
        ws.cell(start, 1, block["objective"] or f"[Objetivo {key}]").alignment = wrap_top

        for i, ind in enumerate(inds):
            r = row + i
            ws.cell(r, 2, ind)
            for col in range(1, 19):
                c = ws.cell(r, col)
                c.border = thin
                if not c.alignment or not c.alignment.wrap_text:
                    c.alignment = wrap_top
        row = end + 1

    return ws


def create_summary_next_steps_table(wb,*,data: pd.DataFrame,sheet_name: str = "summary_next_steps",) -> str:
    # ────── organise data ──────
    spec = {}
    for _, row in data.iterrows():
        t, num, name = row["Element type"], str(row["Number"]), row["Name"]
        if t == "Specific Objective":
            spec.setdefault(num, {"objective": name, "inds": []})
        elif t == "Result indicator":
            key = ".".join(num.split(".")[:2])
            spec.setdefault(key, {"objective": None, "inds": []})
            spec[key]["inds"].append(name)

    order = sorted(spec, key=lambda k: tuple(map(int, k.split("."))))
    for v in spec.values():
        v["inds"].sort()

    # ────── workbook & styles ──────
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]                 # empieza de cero
    ws = wb.create_sheet(sheet_name)
    ws.title = "F. Resumen"
    widths = [
        25, 28, 20, 18, 18, 20, 15,     # A-G
        10, 22, 28, 28,                 # H-K (4 cols + banner filler)
        10, 22, 28, 28, 28              # L-P  (5 cols)
    ]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # palette
    BLUE   = "196E8C"
    GREEN  = "308144"
    GREY   = "E7E6E6"

    bold_white = Font(bold=True, color="FFFFFF")
    bold = Font(bold=True)
    title = Font(bold=True, size=14)
    italic = Font(italic=True, size=11)
    wrap_c = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap_t = Alignment(vertical="top", wrap_text=True)
    thin = Border(*(Side(style="thin"),)*4)

    # ────── rows 1-3  (title, note) ──────
    ws.merge_cells("A1:P1")
    ws["A1"].value, ws["A1"].font = "F. Resumen y próximos pasos", title
    ws.merge_cells("A2:P2")
    ws["A2"].value, ws["A2"].font = (
        "Instrucciones: Resuma los principales desafíos (columnas H a P) "
        "e identifique los próximos pasos (columnas S a U).", italic)
    ws.merge_cells("A3:P3")
    ws["A3"].value = (
        "Nota: Esta sección debe completarse con la información de Desafíos y Acciones potenciales.")

    # ────── row-4  green group banners ──────
    ws.merge_cells("H4:K4")
    c = ws["H4"]; c.value, c.font, c.alignment, c.fill = (
        "Logro del Resultado (Teoría de Cambio)", bold_white, wrap_c,
        PatternFill("solid", fgColor=GREEN))
    ws.merge_cells("L4:P4")
    c = ws["L4"]; c.value, c.font, c.alignment, c.fill = (
        "Medición", bold_white, wrap_c, PatternFill("solid", fgColor=GREEN))

    # ────── row-5  grey sub-banners ──────
    ws.merge_cells("H5:K5")
    s1 = ws["H5"]; s1.value, s1.font, s1.alignment, s1.fill = (
        "Desafíos identificados relacionados al logro de los resultados",
        bold, wrap_c, PatternFill("solid", fgColor=GREY))
    ws.merge_cells("L5:P5")
    s2 = ws["L5"]; s2.value, s2.font, s2.alignment, s2.fill = (
        "Desafíos identificados relacionados con la medición de resultados",
        bold, wrap_c, PatternFill("solid", fgColor=GREY))

    # ────── row-6  column headers ──────
    hdrs = [
        ("A6", "Objetivos Específicos"),
        ("B6", "Indicadores de Resultado"),
        ("C6", "Desagregación"),
        ("D6", "Unidad de Medida"),
        ("E6", "Línea de Base"),
        ("F6", "Año de Línea de Base"),
        ("G6", "Meta"),
        # Logro (H-K)
        ("H6", "Si/No\n(Seleccione una opción)"),
        ("I6", "Tipo de desafío\n(Seleccione una opción)"),
        ("J6", "Explique"),
        ("K6", "Soluciones propuestas"),
        # Medición (L-P)
        ("L6", "Si/No\n(Seleccione una opción)"),
        ("M6", "Tipo de desafío\n(Seleccione una opción)"),
        ("N6", "Explique"),
        ("O6", "¿Se miden todas las dimensiones del Objetivo Específico?"),
        ("P6", "Soluciones propuestas"),
    ]
    for cell, txt in hdrs:
        if cell[0] in "ABCDEFG":  # blue headers
            col_fill, font = BLUE, bold_white
        else:                     # grey challenge headers
            col_fill, font = GREY, bold
        c = ws[cell]; c.value, c.font, c.alignment, c.fill = (
            txt, font, wrap_c, PatternFill("solid", fgColor=col_fill))

    # ────── data rows ──────
    row = 7
    for key in order:
        block, inds = spec[key], spec[key]["inds"]
        start, end = row, row + len(inds) - 1
        ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
        ws.cell(start, 1, block["objective"] or f"[Objetivo {key}]").alignment = wrap_t

        for i, ind in enumerate(inds):
            r = row + i
            ws.cell(r, 2, ind)
            # borders for inner table (A-P) only
            for col in range(1, 17):
                ws.cell(r, col).border = thin
                if not ws.cell(r, col).alignment:  # keep existing for A
                    ws.cell(r, col).alignment = wrap_t
        row = end + 1

    return ws


def create_summary_next_steps_table(wb, *,
    data: pd.DataFrame,
    sheet_name : str =  "create_summary_next_steps_table"
) -> str:
    # ────── 1. Organizar datos ──────
    spec = {}
    for _, row in data.iterrows():
        et, num, name = row["Element type"], str(row["Number"]), row["Name"]
        if et == "Specific Objective":
            spec.setdefault(num, {"objective": name, "inds": []})
        elif et == "Result indicator":
            spec.setdefault(".".join(num.split(".")[:2]), {"objective": None, "inds": []})
            spec[".".join(num.split(".")[:2])]["inds"].append(name)

    order = sorted(spec, key=lambda k: tuple(map(int, k.split("."))))
    for v in spec.values():
        v["inds"].sort()

    # ────── 2. Crear workbook (solo uno) ──────
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]                 # empieza de cero
    ws = wb.create_sheet(sheet_name)
    ws.title = "F. Resumen"

    col_widths = [25, 28, 20, 18, 18, 20, 15, 10, 22, 28, 28, 10, 22, 28, 28, 28]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # Colores
    BLUE, GREEN, GREY = "196E8C", "308144", "E7E6E6"

    # Fuentes / alineación
    f_white = Font(bold=True, color="FFFFFF")
    f_bold  = Font(bold=True)
    title   = Font(bold=True, size=14)
    italic  = Font(italic=True, size=11)
    wrap_c  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap_t  = Alignment(vertical="top", wrap_text=True)
    thin    = Border(*(Side(style="thin"),)*4)

    # ────── 3. Filas 1-3 ──────
    ws.merge_cells("A1:P1"); ws["A1"].value, ws["A1"].font = "F. Resumen y próximos pasos", title
    ws.merge_cells("A2:P2"); ws["A2"].value, ws["A2"].font = (
        "Instrucciones: Resuma los principales desafíos (columnas H-P) e identifique los próximos pasos.",
        italic)
    ws.merge_cells("A3:P3"); ws["A3"].value = (
        "Nota: Esta sección debe completarse con la información de Desafíos y Acciones potenciales.")

    # ────── 4. Banners (filas 4-5) ──────
    ws.merge_cells("H4:K4"); c = ws["H4"]
    c.value, c.font, c.alignment, c.fill = "Logro del Resultado (Teoría de Cambio)", f_white, wrap_c, PatternFill("solid", fgColor=GREEN)
    ws.merge_cells("L4:P4"); c = ws["L4"]
    c.value, c.font, c.alignment, c.fill = "Medición", f_white, wrap_c, PatternFill("solid", fgColor=GREEN)

    ws.merge_cells("H5:K5"); c = ws["H5"]
    c.value, c.font, c.alignment, c.fill = (
        "Desafíos identificados relacionados al logro de los resultados", f_bold, wrap_c, PatternFill("solid", fgColor=GREY))
    ws.merge_cells("L5:P5"); c = ws["L5"]
    c.value, c.font, c.alignment, c.fill = (
        "Desafíos identificados relacionados con la medición de resultados", f_bold, wrap_c, PatternFill("solid", fgColor=GREY))

    # ────── 5. Encabezados fila 6 ──────
    heads = [
        ("A6", "Objetivos Específicos", BLUE, f_white),
        ("B6", "Indicadores de Resultado", BLUE, f_white),
        ("C6", "Desagregación", BLUE, f_white),
        ("D6", "Unidad de Medida", BLUE, f_white),
        ("E6", "Línea de Base", BLUE, f_white),
        ("F6", "Año de Línea de Base", BLUE, f_white),
        ("G6", "Meta", BLUE, f_white),
        ("H6", "Si/No\n(Seleccione una opción)", GREY, f_bold),
        ("I6", "Tipo de desafío\n(Seleccione una opción)", GREY, f_bold),
        ("J6", "Explique", GREY, f_bold),
        ("K6", "Soluciones propuestas", GREY, f_bold),
        ("L6", "Si/No\n(Seleccione una opción)", GREY, f_bold),
        ("M6", "Tipo de desafío\n(Seleccione una opción)", GREY, f_bold),
        ("N6", "Explique", GREY, f_bold),
        ("O6", "¿Se miden todas las dimensiones del Objetivo Específico?", GREY, f_bold),
        ("P6", "Soluciones propuestas", GREY, f_bold),
    ]
    for cell, txt, color, font in heads:
        c = ws[cell]; c.value, c.font, c.alignment, c.fill = txt, font, wrap_c, PatternFill("solid", fgColor=color)

    # ────── 6. Datos ──────
    row = 7
    for key in order:
        bloc, inds = spec[key], spec[key]["inds"]
        ws.merge_cells(start_row=row, start_column=1, end_row=row+len(inds)-1, end_column=1)
        ws.cell(row, 1, bloc["objective"] or f"[Objetivo {key}]").alignment = wrap_t
        for ind in inds:
            ws.cell(row, 2, ind)
            for col in range(1, 17):
                ws.cell(row, col).border = thin
                if ws.cell(row, col).alignment is None:
                    ws.cell(row, col).alignment = wrap_t
            row += 1
    last = row - 1

    # ────── 7. Listas desplegables ──────
    # Sí / No
    yesno = DataValidation(type="list", formula1='"Sí,No"'); ws.add_data_validation(yesno)
    yesno.add(f"H7:H{last}"); yesno.add(f"L7:L{last}")

    # Tipo de desafío (Logro)
    opts_logro = ('"a) Productos no completados o desactivados (no suficientes para alcanzar el resultado),'
                  'b) Productos se completan pero se requiere de más tiempo para alcanzar resultados,'
                  'c) Productos se completan pero no suficientes para alcanzar el resultado,'
                  'd) Otro"')
    dv_logro = DataValidation(type="list", formula1=opts_logro); ws.add_data_validation(dv_logro)
    dv_logro.add(f"I7:I{last}")

    # Tipo de desafío (Medición)
    opts_med = ('"a) El indicador y/o su metodología de cálculo no están adecuadamente definidos,'
                'b) Dificultad para acceder a datos de fuentes existentes,'
                'c) Dificultad para recopilar nuevos datos por el proyecto,'
                'd) Otro"')
    dv_med = DataValidation(type="list", formula1=opts_med); ws.add_data_validation(dv_med)
    dv_med.add(f"M7:M{last}")

    return ws



def create_theory_of_change_table(wb,*,results_df: pd.DataFrame,components_df: pd.DataFrame, sheet_name ="theory_of_change") -> str:
    """
    • Tabla A (cambios en los productos) columnas A-O
        - A = Name (solo Element type == 'Solution')
        - B = ID   (solo Element type == 'Solution')
        - Fórmulas automáticas:
              *  L =IF(SUM(G#:K#)>0,1,0)
              *  N =IF(MAX(R#:AZ#)=2,1,0)
    • Tabla B (matriz horizontal de contribución) desde R
    • Todos los anchos de columna = 15 (caracteres)
    • Todo alineado a la izquierda, salvo encabezados rotados (centrados)
    """

    # ── 1 · DATOS DE RESULTADOS (Objetivos / Indicadores) ─────────
    spec: dict[str, dict] = {}
    for _, r in results_df.iterrows():
        e_type, num, name = r["Element type"], str(r["Number"]), r["Name"]
        if e_type == "Specific Objective":
            spec.setdefault(num, {"objective": name, "inds": []})
        elif e_type == "Result indicator":
            key = ".".join(num.split(".")[:2])
            spec.setdefault(key, {"objective": None, "inds": []})
            spec[key]["inds"].append(name)

    ordered_specs = sorted(spec, key=lambda x: tuple(map(int, x.split("."))))
    for v in spec.values():
        v["inds"].sort()

    # ── 2 · COMPONENTES TIPO SOLUTION ─────────────────────────────
    sol_df = components_df[components_df["Element type"].str.lower() == "solution"]

    # ── 3 · LIBRO Y ESTILOS BÁSICOS ───────────────────────────────
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]                 # empieza de cero
    ws = wb.create_sheet(sheet_name)
    ws.title = "D. Teoría de Cambio"

    # ancho uniforme 15 para “muchas” columnas
    for col_idx in range(1, 200):
        ws.column_dimensions[get_column_letter(col_idx)].width = 19

    # paleta
    BLUE, GREEN, LGREEN = "196E8C", "308144", "ACCDB4"
    GREY, HEADER_GREY = "A6A6A6", "E7E6E6"
    CYAN, CREME = "00B0F0", "F2F2CB"

    # estilos
    bold_white = Font(bold=True, color="FFFFFF")
    bold = Font(bold=True)
    title_f = Font(bold=True, size=14)
    italic_f = Font(italic=True, size=11)
    left_mid = Alignment(horizontal="left", vertical="center", wrap_text=True)
    left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
    vertical_hdr = Alignment(horizontal="center", vertical="center", wrap_text=True, text_rotation=90)
    thin = Border(*(Side(style="thin"),) * 4)

    # ── 4 · CABECERA GENERAL ──────────────────────────────────────
    ws.merge_cells("A1:O1")
    ws["A1"].value, ws["A1"].font, ws["A1"].alignment = "D. Teoría de Cambio", title_f, left_mid
    ws.merge_cells("A2:O2")
    ws["A2"].value, ws["A2"].font, ws["A2"].alignment = (
        "Instrucciones: Por favor llene las secciones A y B (secciones en verde).",
        italic_f,
        left_mid,
    )

    # ── 5 · TABLA A  (A-O) ────────────────────────────────────────
    ws.merge_cells("G4:K4")
    ban_a = ws["G4"]
    ban_a.value = ("A. Cambios en los Productos\n"
                   "Rellene la celda con el valor 1 si alguna de las siguientes opciones aplica a cada producto.")
    ban_a.font, ban_a.alignment, ban_a.fill = bold_white, left_mid, PatternFill("solid", fgColor=GREEN)
    ws.row_dimensions[4].height = 45

    hdr_a = [
        ("A5", "Declaración de componentes", BLUE,  bold_white),
        ("B5", "ID Componente",              BLUE,  bold_white),
        ("C5", "ID Producto",                BLUE,  bold_white),
        ("D5", "Definición del Producto",    BLUE,  bold_white),
        ("E5", "Producto Desactivado",       GREY,  bold),
        ("F5", "Advertencia",                BLUE,  bold_white),
        ("G5", "Cancelado o Desactivado",         LGREEN, bold),
        ("H5", "Retrasado",                       LGREEN, bold),
        ("I5", "Cambio en el Alcance Financiero", LGREEN, bold),
        ("J5", "Cambio en el Alcance Físico",     LGREEN, bold),
        ("K5", "Nuevo producto",                 LGREEN, bold),
        ("L5", "El Producto ha sufrido cambios",  GREY,   bold),
        ("M5", "Para productos con cambios,\nexplique las causas", CREME, bold),
        ("N5", "Productos gatilladores\npara logro de resultados", GREY, bold),
        ("O5", "Para los productos gatilladores,\nidentifique los principales supuestos\npara el logro de resultados",
               CYAN, bold),
    ]

    for cell, txt, color, fnt in hdr_a:
        c = ws[cell]
        c.value, c.font = txt, fnt
        # rotar los estrictamente verticales
        c.alignment = vertical_hdr if cell in ("C5", "E5", "F5") else left_mid
        c.fill = PatternFill("solid", fgColor=color)

    # escribir componentes Solution (A-B) y bordear fila
    row = 6
    for _, comp in sol_df.iterrows():
        ws.cell(row, 1, comp["Name"]).alignment = left_top
        ws.cell(row, 2, comp["ID"]).alignment = left_top
        for col in range(1, 16):       # A-O
            ws.cell(row, col).border = thin
        # ===<<  SECCIÓN FÓRMULAS  >>========================================
        # L: 1 si algún cambio marcado en G-K
        ws.cell(row, 12).value = f"=IF(SUM(G{row}:K{row})>0,1,0)"
        # N: 1 si en la matriz horizontal (R:AZ) existe un 2
        ws.cell(row, 14).value = f"=IF(MAX(R{row}:AZ{row})=2,1,0)"
        # ===================================================================
        row += 1

    # ── 6 · TABLA B  (R en adelante) ──────────────────────────────
    start_col = 18  # R
    # calcular última columna necesaria
    last_col = start_col
    for k in ordered_specs:
        last_col += 1 + len(spec[k]["inds"])

    ws.merge_cells(start_row=4, start_column=start_col, end_row=4, end_column=last_col - 1)
    ban_b = ws.cell(4, start_col)
    ban_b.value = ("B. Teoría de Cambio\n"
                   "Para cada Objetivo Específico e Indicador marque:\n"
                   "1 – si el producto contribuye a su logro\n"
                   "2 – si el producto contribuye y es necesario para su logro")
    ban_b.font, ban_b.alignment, ban_b.fill = bold_white, left_mid, PatternFill("solid", fgColor=GREEN)
    ws.row_dimensions[4].height = 60

    # Fila 5: nombres reales de Objetivos + indicadores
    col = start_col
    for key in ordered_specs:
        obj_name = spec[key]["objective"] or f"[Objetivo {key}]"
        obj_cell = ws.cell(5, col, obj_name)
        obj_cell.font, obj_cell.alignment, obj_cell.fill = bold, left_mid, PatternFill("solid", fgColor=HEADER_GREY)
        col += 1
        for ind in spec[key]["inds"]:
            c_ind = ws.cell(5, col, ind)
            c_ind.font, c_ind.alignment, c_ind.fill = bold, left_mid, PatternFill("solid", fgColor=HEADER_GREY)
            col += 1

    # bordes y alineación (filas 6-15)
    for r in range(6, 16):
        for c in range(start_col, col):
            ws.cell(r, c).border = thin
            ws.cell(r, c).alignment = left_mid

    return ws


